import tkinter as tk
from tkinter import ttk
from tkinter import Menu
from datetime import datetime
import openpyxl
import os

class Application(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Navegação entre Telas")
        self.geometry("600x500")  # Definindo um tamanho inicial para a janela

        # Criar um menu
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        # Criar o menu de Telas
        tela_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Telas", menu=tela_menu)

        # Adicionar as opções do menu
        tela_menu.add_command(label="AlfaceBom", command=self.show_alface_bom)
        tela_menu.add_command(label="Análises", command=self.show_analises)
        tela_menu.add_command(label="Dados", command=self.show_dados)

        # Inicializar a tela atual
        self.current_screen = None
        self.show_alface_bom()

    def show_alface_bom(self):
        self.switch_screen_alface_bom()

    def show_analises(self):
        self.switch_screen("Bem-vindo à tela Análises!")

    def show_dados(self):
        self.switch_screen("Bem-vindo à tela Dados!")

    def switch_screen(self, content):
        if self.current_screen:
            self.current_screen.destroy()

        self.current_screen = tk.Label(self, text=content)
        self.current_screen.pack(fill="both", expand=True)

    def switch_screen_alface_bom(self):
        if self.current_screen:
            self.current_screen.destroy()

        self.current_screen = AlfaceBomScreen(self)
        self.current_screen.pack(fill="both", expand=True)

class AlfaceBomScreen(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master

        self.vendedor_var = tk.StringVar()
        self.quantidade_var = tk.StringVar()
        self.data_hora_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.cliente_var = tk.StringVar()

        self.create_form()
        self.create_table()
        self.create_console()

        self.load_data_from_excel()

    def create_form(self):
        form_frame = tk.Frame(self)
        form_frame.pack(padx=20, pady=20, fill="x")

        tk.Label(form_frame, text="Vendedor:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(form_frame, textvariable=self.vendedor_var).grid(row=0, column=1, padx=5, pady=5)

        tk.Label(form_frame, text="Quantidade:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        tk.Entry(form_frame, textvariable=self.quantidade_var).grid(row=0, column=3, padx=5, pady=5)

        tk.Label(form_frame, text="Data e Hora:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(form_frame, textvariable=self.data_hora_var).grid(row=1, column=1, padx=5, pady=5)

        tk.Label(form_frame, text="Cliente:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        tk.Entry(form_frame, textvariable=self.cliente_var).grid(row=1, column=3, padx=5, pady=5)

        tk.Button(form_frame, text="Adicionar", command=self.add_entry).grid(row=2, columnspan=4, pady=10)

    def create_table(self):
        table_frame = tk.Frame(self)
        table_frame.pack(padx=20, fill="both", expand=True)

        self.table = ttk.Treeview(table_frame, columns=("Vendedor", "Quantidade", "Data e Hora", "Cliente"), show="headings")
        self.table.heading("Vendedor", text="Vendedor")
        self.table.heading("Quantidade", text="Quantidade")
        self.table.heading("Data e Hora", text="Data e Hora")
        self.table.heading("Cliente", text="Cliente")
        self.table.bind("<Button-3>", self.show_popup_menu)

        self.table.pack(fill="both", expand=True)
        self.update_table()

    def create_console(self):
        self.console = tk.Text(self, wrap="word", height=6)
        self.console.pack(padx=20, pady=(0, 20), fill="both", expand=True)

    def log_to_console(self, message):
        self.console.insert("end", message + "\n")
        self.console.see("end")

    def add_entry(self):
        vendedor = self.vendedor_var.get()
        quantidade = self.quantidade_var.get()
        data_hora = self.data_hora_var.get()
        cliente = self.cliente_var.get()

        if vendedor and quantidade and data_hora and cliente:
            self.table.insert("", "end", values=(vendedor, quantidade, data_hora, cliente))
            self.clear_form()
            self.save_to_excel()
            self.log_to_console("Nova entrada adicionada")
        else:
            self.log_to_console("Campos incompletos. Não foi possível adicionar a entrada.")

    def clear_form(self):
        self.vendedor_var.set("")
        self.quantidade_var.set("")
        self.data_hora_var.set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.cliente_var.set("")

    def update_table(self):
        for row in self.table.get_children():
            self.table.delete(row)

    def show_popup_menu(self, event):
        item = self.table.identify_row(event.y)
        if item:
            popup_menu = Menu(self, tearoff=0)
            popup_menu.add_command(label="Editar", command=lambda: self.edit_entry(item))
            popup_menu.add_command(label="Excluir", command=lambda: self.delete_entry(item))
            popup_menu.tk_popup(event.x_root, event.y_root)

    def edit_entry(self, item):
        values = self.table.item(item, "values")
        if values:
            self.vendedor_var.set(values[0])
            self.quantidade_var.set(values[1])
            self.data_hora_var.set(values[2])
            self.cliente_var.set(values[3])
            self.table.delete(item)
            self.save_to_excel()
            self.log_to_console("Entrada editada")

    def delete_entry(self, item):
        self.table.delete(item)
        self.save_to_excel()
        self.log_to_console("Entrada excluída")

    def save_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Vendedor", "Quantidade", "Data e Hora", "Cliente"])
        
        for item in self.table.get_children():
            values = self.table.item(item, "values")
            ws.append(values)
        
        filename = "dados.xlsx"
        wb.save(filename)

    def load_data_from_excel(self):
        filename = "dados.xlsx"
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.table.insert("", "end", values=row)

if __name__ == "__main__":
    app = Application()
    app.mainloop()